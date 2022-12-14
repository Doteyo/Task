import csv
import codecs
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, Border, Side
from matplotlib import pyplot as plt
import numpy as np
import re
import sys

import prettytable
from prettytable import PrettyTable


class Report:
    year_prof_sal = year_prof_vacs = border = ""

    def __init__(self, year_sal, year_vacs, year_prof_sal, year_prof_vacs, city_sal, city_part):
        self.year_prof_sal = year_prof_sal
        self.year_prof_vacs = year_prof_vacs
        self.years_l = [*year_sal.keys()]
        self.year_sal_l = [*year_sal.values()]
        self.year_prof_sal_l = [*fill_gaps(year_prof_sal, year_sal, 0).values()]
        self.year_vacs_l = [*year_vacs.values()]
        self.year_prof_vacs_l = [*fill_gaps(year_prof_vacs, year_sal, 0).values()]
        self.city_A_l = [*city_sal.keys()]
        self.sal_A_l = [*city_sal.values()]
        self.city_B_l = [*city_part.keys()]
        self.part_B_l = [*city_part.values()]

    def print_data(self):
        print(f"Динамика уровня зарплат по годам: {dict(zip(self.years_l, self.year_sal_l))}")
        print(f"Динамика количества вакансий по годам: {dict(zip(self.years_l, self.year_vacs_l))}")
        print(f"Динамика уровня зарплат по годам для выбранной профессии: {self.year_prof_sal}")
        print(f"Динамика количества вакансий по годам для выбранной профессии: {self.year_prof_vacs}")
        print(f"Уровень зарплат по городам (в порядке убывания): {dict(zip(self.city_A_l, self.sal_A_l))}")
        print(f"Доля вакансий по городам (в порядке убывания): {dict(zip(self.city_B_l, self.part_B_l))}")

    def generate_excel(self):
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Статистика по годам"
        ws2 = wb.create_sheet("Статистика по городам")
        ws2.append(['Город', 'Уровень зарплат', "", 'Город', 'Доля вакансий'])
        thin = Side(style="thin", color="FF000000")
        self.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        self.fill_columns(1, 1, ws1, self.twod_array(
            ["Год", "Средняя зарплата", f"Средняя зарплата - {prof_name}", "Количество вакансий",
             f"Количество вакансий - {prof_name}"]))
        self.fill_columns(1, 2, ws1, [self.years_l, self.year_sal_l, self.year_prof_sal_l,
                                      self.year_vacs_l, self.year_prof_vacs_l])
        self.fill_columns(1, 1, ws2, self.twod_array(['Город', 'Уровень зарплат', "", 'Город', 'Доля вакансий']))
        self.fill_columns(1, 2, ws2, [self.city_A_l, self.sal_A_l, [""] * 10, self.city_B_l, self.part_B_l])
        self._format_column_width(ws1, ws2)
        wb.save("report.xlsx")

    def twod_array(self, data: list):
        return list(map(lambda x: [x], data))

    def fill_columns(self, start_col, start_row, sheet, arrays: list):
        max_col_index = len(arrays) - 1
        max_row_index = len(max(arrays, key=len))
        i = 0
        for col in sheet.iter_cols(min_col=start_col, max_col=max_col_index + start_col,
                                   max_row=max_row_index + start_row, min_row=start_row):
            for j in range(len(arrays[i])):
                col[j].value = arrays[i][j]
                if arrays[i][j] != "" and col[j] is not None:
                    if col[j].row == 1:
                        col[j].font = Font(bold=True)
                    if isinstance(arrays[i][j], float):
                        col[j].style = "Percent"
                    col[j].border = self.border
            i += 1

    def _format_column_width(self, sheet1, sheet2):
        self._set_max_width_and_styles(sheet1, len(self.years_l) + 1)
        self._set_max_width_and_styles(sheet2, len(self.city_A_l) + 10)

    def _set_max_width_and_styles(self, sheet, length):
        for col in sheet.iter_cols(max_col=5, min_row=1, max_row=length):
            sheet.column_dimensions[col[0].column_letter].width = \
                len(str(max(col, key=lambda x: len(str(x.value)) if x.value is not None else 0).value)) + 2

    def generate_image(self):
        figure, axis = plt.subplots(2, 2)
        bar_x = np.arange(len(self.years_l))
        axis[0, 0].bar(bar_x - 0.2, self.year_sal_l, 0.4, label="средняя з/п")
        axis[0, 0].bar(bar_x + 0.2, self.year_prof_sal_l, 0.4, label=f"з/п {prof_name}")
        axis[0, 0].set_xticks(bar_x, self.years_l, rotation=90, fontsize=8)
        axis[0, 0].set_title("Уровень зарплат по годам")
        axis[0, 0].legend(fontsize=8)
        axis[0, 0].grid(visible=True, axis="y")
        axis[0, 1].bar(bar_x - 0.2, self.year_vacs_l, 0.4, label="Количество вакансий")
        axis[0, 1].bar(bar_x + 0.2, self.year_prof_vacs_l, 0.4, label=f"Количество вакансий \n{prof_name}")
        axis[0, 1].set_xticks(bar_x, self.years_l, rotation=90, fontsize=8)
        axis[0, 1].set_title("Количество вакансий по годам")
        axis[0, 1].legend(loc="upper left", fontsize=8)
        axis[0, 1].grid(visible=True, axis="y")
        for i in range(len(self.city_A_l)):
            if self.city_A_l[i].count("-") == 1 or self.city_A_l[i].count(" ") == 1:
                self.city_A_l[i] = self.city_A_l[i].replace("-", "-\n", 1).replace(" ", " \n", 1)
        axis[1, 0].barh(np.arange(len(self.city_A_l)), self.sal_A_l, align="edge")
        axis[1, 0].set_yticks(np.arange(len(self.city_A_l)), labels=self.city_A_l, fontsize=6)
        axis[1, 0].invert_yaxis()
        axis[1, 0].grid(visible=True, axis="x")
        axis[1, 0].set_title("Уровень зарплат по городам")
        self.part_B_l = [1 - sum(self.part_B_l)] + self.part_B_l
        self.city_B_l = ["Другие"] + self.city_B_l
        axis[1, 1].pie(self.part_B_l, labels=self.city_B_l, textprops={'fontsize': 6})
        axis[1, 1].set_title("Доля вакансий по городам")
        plt.tight_layout()
        plt.savefig('graph.png')


class ProfKeys:
    def __init__(self, headers):
        self.name = headers.index('name')
        self.salary_from = headers.index('salary_from')
        self.salary_to = headers.index('salary_to')
        self.salary_currency = headers.index('salary_currency')
        self.area_name = headers.index('area_name')
        self.published_at = headers.index('published_at')


currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}

salary_all_years = {}
count_all_vacs = {}
salary_prof_years = {}
count_prof_vacs = {}
salary_city = {}
part_city = {}
salary_city_part = {}
count_city_vacs = {}


def fill(cur: dict, ref: dict, filler):
    if len(cur) == 0:
        for x in ref.keys():
            cur[x] = filler


def fill_gaps(cur: dict, ref: dict, filler):
    temp = cur
    for x in ref.keys():
        if x not in temp:
            temp[x] = filler
    return temp


def year(ls):
    return int(ls[Keys.published_at][0:4])


def sal(*sal_list):
    return currency_to_rub[sal_list[2]] * (float(sal_list[0]) + float(sal_list[1])) / 2


def for_loop_div(key_source: dict, divide: dict, action):
    for x in key_source.keys():
        divide[x] = action(divide[x], key_source[x])


def csv_reader(file_name):
    file = codecs.open(file_name, 'r', 'utf_8_sig')
    reader = csv.reader(file)
    data = list(reader)
    file.close()
    return data


def addToDict(key_val, dict, val):
    if key_val in dict.keys():
        dict[key_val] += val
    else:
        dict[key_val] = val


def create_dicts(data, name):
    number = 0
    for line in data:
        if all(line):
            number += 1
            addToDict(year(line), salary_all_years,
                      sal(line[Keys.salary_from], line[Keys.salary_to], line[Keys.salary_currency]))
            addToDict(line[Keys.area_name], salary_city,
                      sal(line[Keys.salary_from], line[Keys.salary_to], line[Keys.salary_currency]))
            addToDict(year(line), count_all_vacs, 1)
            addToDict(line[Keys.area_name], count_city_vacs, 1)
            if name in line[Keys.name]:
                addToDict(year(line), salary_prof_years,
                          sal(line[Keys.salary_from], line[Keys.salary_to], line[Keys.salary_currency]))
                addToDict(year(line), count_prof_vacs, 1)
    for_loop_div(count_all_vacs, salary_all_years, lambda x, y: int(x / y))
    for_loop_div(count_prof_vacs, salary_prof_years, lambda x, y: int(x / y))
    return number


def calculate_part_city(num):
    for x in count_city_vacs.keys():
        calc_num = Decimal(count_city_vacs[x] / num).quantize(Decimal("1.0000"))
        if calc_num >= 0.01:
            part_city[x] = calc_num.__float__()
            salary_city_part[x] = int(salary_city[x] / count_city_vacs[x])


def alphabetic_sort(ls: list):
    temp_list = []
    for i in range(1, len(ls)):
        if ls[i][1] == ls[i - 1][1]:
            temp_list.append(ls[i])
            if ls[i - 1] not in temp_list:
                temp_list.append(ls[i - 1])
        elif len(temp_list) > 1 and ls[i][1] != ls[i - 1][1]:
            ls[ls.index(temp_list[0]) - 1:i] = sorted(temp_list, key=lambda f: f[0])
            temp_list.clear()

def printDict(d):
    for key, value in d.items():
        print(key + ':', value)


def csv_filter(reader, list_naming):
    temp_list = []
    for line in reader:
        filtered_line = [re.sub(r"<[^>]+>", "", b, flags=re.S) for b in line]
        if all(line) and len(filtered_line) == len(list_naming):
            temp_list.append(
                [re.sub(r'\s+', " ", ', '.join(
                    [f.strip() for f in g.split("\n")])) for g in filtered_line])
    dicts_list = [dict(zip(list_naming, line)) for line in temp_list]
    return dicts_list


def print_vacancies(data_vacancies, dic_naming):
    counter = 1
    table = PrettyTable()
    table.align = 'l'
    table.hrules = prettytable.ALL
    for el in data_vacancies:
        formatter(el)
        for key in dic_naming:
            if key in el:
                el[dic_naming[key]] = el.pop(key)
        if len(table.field_names) == 0:
            table.field_names = ["№"] + [x for x in el.keys()]
            table._max_width = dict(zip(table.field_names, len(table.field_names) * [20]))
        table.add_row([counter] + [r for r in el.values()])
        counter += 1
    print(table)


def formatter(row):
    trim = lambda x: x[0:100] + "..." if len(x) > 100 else x
    row['experience_id'] = replacement_dic[row['experience_id']]
    row["premium"] = row["premium"].replace("False", "Нет").replace("True", "Да")
    gross = lambda x: "Без вычета налогов" if x == "True" else "С вычетом налогов"
    row["salary_range"] = f"{'{:,.0f}'.format(float(row.pop('salary_from'))).replace(',', ' ')} " \
                          f"- {'{:,.0f}'.format(float(row.pop('salary_to'))).replace(',', ' ')} " \
                          f"({replacement_dic[row.pop('salary_currency')]}) " \
                          f"({gross(row.pop('salary_gross'))})"
    row["publish_day"] = ".".join(row.pop("published_at")[:10].split("-")[::-1])
    row["key_skills"] = row["key_skills"].replace(', ', '\n')
    for key in row:
        row[key] = trim(row[key])

choice = input("Вакансии или Статистика: ")
while choice not in ["Вакансии", "Статистика"]:
    choice = input("Вакансии или Статистика: ")

if choice == "Статистика":
    file_name = input("Введите название файла: ")
    prof_name = input("Введите название профессии: ")

    data = csv_reader(file_name)
    Keys = ProfKeys(data.pop(0))
    number_of_vacs = create_dicts(data, prof_name)
    calculate_part_city(number_of_vacs)
    city_part_vacs = sorted(part_city.items(), key=lambda x: x[1], reverse=True)
    city_salary = sorted(salary_city_part.items(), key=lambda x: x[1], reverse=True)
    alphabetic_sort(city_salary)
    fill(salary_prof_years, salary_all_years, 0)
    fill(count_prof_vacs, count_all_vacs, 0)
    report = Report(salary_all_years, count_all_vacs, salary_prof_years, count_prof_vacs, dict(city_salary[0:10]),
                    dict(city_part_vacs[0:10]))
    report.print_data()
    report.generate_image()
else:
    temp = csv_reader(input())

    replacement_dic = dict(name="Название", description="Описание", key_skills="Навыки", experience_id="Опыт работы",
                           premium="Премиум-вакансия", employer_name="Компания", salary_from="Нижняя граница вилки оклада",
                           salary_to="Верхняя граница вилки оклада", salary_range="Оклад",
                           salary_gross="Оклад указан до вычета налогов", salary_currency="Идентификатор валюты оклада",
                           area_name="Название региона", published_at="Дата и время публикации вакансии",
                           publish_day="Дата публикации вакансии", AZN="Манаты", BYR="Белорусские рубли", EUR="Евро",
                           GEL="Грузинский лари", KGS="Киргизский сом", KZT="Тенге", RUR="Рубли", UAH="Гривны",
                           USD="Доллары", UZS="Узбекский сум", noExperience="Нет опыта", between1And3="От 1 года до 3 лет",
                           between3And6="От 3 до 6 лет", moreThan6="Более 6 лет")

    if len(temp) == 0:
        print("Пустой файл")
        sys.exit()

    dateFiltered = csv_filter(temp[1:], temp[0])
    if len(dateFiltered) == 0:
        print("Нет данных")
    else:
        print_vacancies(dateFiltered, replacement_dic)






