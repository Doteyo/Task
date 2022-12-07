import csv
import codecs
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, Border, Side
from matplotlib import pyplot as plt
import numpy as np
import doctest


class Report:
    """Класс для сбора информации и генерации по ней изображения графиков или таблицу xlsx


    Attributes:
        year_prof_sal (dict(int,int)): Словарь содержащий динамику уровня зп по годам для выбранной профессии
        year_prof_vacs (dict(int,int)): Словарь содержащий кол-во вакансий по годам для выбранной профессии
        years_l (list(int)): Список годов для общей статистики
        year_sal_l (list(int)): Список зп для общей статистики
        year_prof_sal_l (list(int)): Список зп по годам для выбранной профессии с нулями, где профессия не появлялась
        year_vacs_l (list(str)): Список кол-ва вакансий по годам
        year_prof_vacs_l (list(str)): Список кол-ва вакансий по годам для выбранной профессии
        city_A_l (list(str)): Список городов для средней зп
        sal_A_l (list(int)): Список средних зп по городам
        city_B_l (list(str)): Список городов для долей от общего кол-ва вакансий
        part_B_l (list(float)): Список соотношений от общего кол-ва вакансий
    """
    year_prof_sal = year_prof_vacs = border = ""

    def __init__(self, year_sal, year_vacs, year_prof_sal, year_prof_vacs, city_sal, city_part):
        """Инициализирует Report,


        Args:
            :param year_sal: Словарь с средней зп по годам
            :param year_vacs: Словарь с кол-вом профессий по годам
            :param year_prof_sal: Словарь с средней зп по годам для выбранной профессии
            :param year_prof_vacs: Словарь с кол-вом профессий по годам для выбранной профессии
            :param city_sal: Словарь с уровнем зп по городам
            :param city_part: Словарь с долей вакансий по городам
        """
        self.year_prof_sal = year_prof_sal
        self.year_prof_vacs = year_prof_vacs
        self.years_l = [*year_sal.keys()]
        self.year_sal_l = [*year_sal.values()]
        self.year_prof_sal_l = [*fill_gaps(year_prof_sal, year_sal, 0).values()]
        self.year_vacs_l = [*year_vacs.values()]
        self.year_prof_vacs_l = [*fill_gaps(year_prof_vacs, year_sal, 0).values()]
        self.city_A_l = [*city_sal.keys()]
        self.sal_A_l = [*city_sal.values()]
        self.city_B_l = [*city_part.keys()]
        self.part_B_l = [*city_part.values()]

    def print_data(self):
        """Выводит значения в консоль


        """

        print(f"Динамика уровня зарплат по годам: {dict(zip(self.years_l, self.year_sal_l))}")
        print(f"Динамика количества вакансий по годам: {dict(zip(self.years_l, self.year_vacs_l))}")
        print(f"Динамика уровня зарплат по годам для выбранной профессии: {self.year_prof_sal}")
        print(f"Динамика количества вакансий по годам для выбранной профессии: {self.year_prof_vacs}")
        print(f"Уровень зарплат по городам (в порядке убывания): {dict(zip(self.city_A_l, self.sal_A_l))}")
        print(f"Доля вакансий по городам (в порядке убывания): {dict(zip(self.city_B_l, self.part_B_l))}")

    def generate_excel(self):
        """Создаёт report.xlsx файл с предоставленными данными


        """

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Статистика по годам"
        ws2 = wb.create_sheet("Статистика по городам")
        ws2.append(['Город', 'Уровень зарплат', "", 'Город', 'Доля вакансий'])
        thin = Side(style="thin", color="FF000000")
        self.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        self.fill_columns(1, 1, ws1, self.twod_array(
            ["Год", "Средняя зарплата", f"Средняя зарплата - {prof_name}", "Количество вакансий",
             f"Количество вакансий - {prof_name}"]))
        self.fill_columns(1, 2, ws1, [self.years_l, self.year_sal_l, self.year_prof_sal_l,
                                      self.year_vacs_l, self.year_prof_vacs_l])
        self.fill_columns(1, 1, ws2, self.twod_array(['Город', 'Уровень зарплат', "", 'Город', 'Доля вакансий']))
        self.fill_columns(1, 2, ws2, [self.city_A_l, self.sal_A_l, [""] * 10, self.city_B_l, self.part_B_l])
        self._format_column_width(ws1, ws2)
        wb.save("report.xlsx")

    def twod_array(self, data: list):
        """ Переделывает каждый элемент листа в отделынй список


        :param data: Список с данными
        :return: list(list()) Список с списками
        """

        return list(map(lambda x: [x], data))

    def fill_columns(self, start_col, start_row, sheet, arrays: list):
        """ Заполняет столбцы таблицы по данным


        :param start_col: Индекс начального столбца
        :param start_row: Индекс начальной строки
        :param sheet: Страница excel для записис
        :param arrays: Список с списками, где каждый список это столбец
        """

        max_col_index = len(arrays) - 1
        max_row_index = len(max(arrays, key=len))
        i = 0
        for col in sheet.iter_cols(min_col=start_col, max_col=max_col_index + start_col,
                                   max_row=max_row_index + start_row, min_row=start_row):
            for j in range(len(arrays[i])):
                col[j].value = arrays[i][j]
                if arrays[i][j] != "" and col[j] is not None:
                    if col[j].row == 1:
                        col[j].font = Font(bold=True)
                    if isinstance(arrays[i][j], float):
                        col[j].style = "Percent"
                    col[j].border = self.border
            i += 1

    def _format_column_width(self, sheet1, sheet2):
        """ Выставляет ширину столбцам

        :param sheet1: Лист1
        :param sheet2: Лист2
        """
        self._set_max_width_and_styles(sheet1, len(self.years_l) + 1)
        self._set_max_width_and_styles(sheet2, len(self.city_A_l) + 10)

    def _set_max_width_and_styles(self, sheet, length):
        """ Устанавливает ширину столбца по максимальн длинной строке в столбце

        :param sheet: Лист excel
        :param length: Глубина поиска в столбце
        """
        for col in sheet.iter_cols(max_col=5, min_row=1, max_row=length):
            sheet.column_dimensions[col[0].column_letter].width = \
                len(str(max(col, key=lambda x: len(str(x.value)) if x.value is not None else 0).value)) + 2

    def generate_image(self):
        """ Создаёт изображение с графиками по данным
        """
        figure, axis = plt.subplots(2, 2)
        bar_x = np.arange(len(self.years_l))
        axis[0, 0].bar(bar_x - 0.2, self.year_sal_l, 0.4, label="средняя з/п")
        axis[0, 0].bar(bar_x + 0.2, self.year_prof_sal_l, 0.4, label=f"з/п {prof_name}")
        axis[0, 0].set_xticks(bar_x, self.years_l, rotation=90, fontsize=8)
        axis[0, 0].set_title("Уровень зарплат по годам")
        axis[0, 0].legend(fontsize=8)
        axis[0, 0].grid(visible=True, axis="y")
        axis[0, 1].bar(bar_x - 0.2, self.year_vacs_l, 0.4, label="Количество вакансий")
        axis[0, 1].bar(bar_x + 0.2, self.year_prof_vacs_l, 0.4, label=f"Количество вакансий \n{prof_name}")
        axis[0, 1].set_xticks(bar_x, self.years_l, rotation=90, fontsize=8)
        axis[0, 1].set_title("Количество вакансий по годам")
        axis[0, 1].legend(loc="upper left", fontsize=8)
        axis[0, 1].grid(visible=True, axis="y")
        for i in range(len(self.city_A_l)):
            if self.city_A_l[i].count("-") == 1 or self.city_A_l[i].count(" ") == 1:
                self.city_A_l[i] = self.city_A_l[i].replace("-", "-\n", 1).replace(" ", " \n", 1)
        axis[1, 0].barh(np.arange(len(self.city_A_l)), self.sal_A_l, align="edge")
        axis[1, 0].set_yticks(np.arange(len(self.city_A_l)), labels=self.city_A_l, fontsize=6)
        axis[1, 0].invert_yaxis()
        axis[1, 0].grid(visible=True, axis="x")
        axis[1, 0].set_title("Уровень зарплат по городам")
        self.part_B_l = [1 - sum(self.part_B_l)] + self.part_B_l
        self.city_B_l = ["Другие"] + self.city_B_l
        axis[1, 1].pie(self.part_B_l, labels=self.city_B_l, textprops={'fontsize': 6})
        axis[1, 1].set_title("Доля вакансий по городам")
        plt.tight_layout()
        plt.savefig('graph.png')


class ProfKeys:
    """ Класс для хранения индексов нужных строк
        Attributes:
            name (str): Название
            salary_from (int): Мин зп
            salary_to (int): Макс зп
            salary_currency (str): Валюта зп
            area_name (str): Место вакансии
            published_at (str): Дата публикации
    """
    def __init__(self, headers):
        """ Инициализирует класс с данными

        :param headers: Список с заголовками

        >>> type(ProfKeys(['name','salary_from','salary_to','salary_currency','area_name','published_at'])).__name__
        'ProfKeys'

        >>> ProfKeys(['name','salary_from','salary_to','salary_currency','area_name','published_at']).area_name
        4

        >>> ProfKeys(['name','salary_from','salary_to', 'sos','salary_currency','area_name','published_at']).area_name
        5

        """
        self.name = headers.index('name')
        self.salary_from = headers.index('salary_from')
        self.salary_to = headers.index('salary_to')
        self.salary_currency = headers.index('salary_currency')
        self.area_name = headers.index('area_name')
        self.published_at = headers.index('published_at')


currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}

salary_all_years = {}
count_all_vacs = {}
salary_prof_years = {}
count_prof_vacs = {}
salary_city = {}
part_city = {}
salary_city_part = {}
count_city_vacs = {}


def fill(cur: dict, ref: dict, filler):
    """ Заполняет словарь ключами из второго словаря с введённым значением


    :param cur: Текущий словарь
    :param ref: Словарь с ключами
    :param filler: Значение
    """
    if len(cur) == 0:
        for x in ref.keys():
            cur[x] = filler


def fill_gaps(cur: dict, ref: dict, filler):
    """ Заполняет словарь значением filler если в temp нет ключа из ref

    :param cur: Заполняемый словарь
    :param ref: Эталонный словарь
    :param filler: Значение для заполнения
    :return: dict Заполненый словарь

    >>> fill_gaps({},{"a": 1,'b': 2},0)
    {'a': 0, 'b': 0}

    >>> fill_gaps({'a': 5},{"a": 1,'b': 2},'yes')
    {'a': 5, 'b': 'yes'}
    """
    temp = cur
    for x in ref.keys():
        if x not in temp:
            temp[x] = filler
    return temp


def year(ls):
    """ Извлекает год из списка

    :param ls: Список
    :return: int Год из списка

    >>> Keys.published_at = 2
    >>> year(["This",'is','5'])
    5

    >>> Keys.published_at = 2
    >>> year(["This",'is','12345'])
    1234

    """
    return int(ls[Keys.published_at][0:4])


def sal(*sal_list):
    """ Расчитывает среднюю зарплату


    :param sal_list:  Список параметров для расчёта зп
    :return: float Средняя зп

    >>> sal(20000,50000,'RUR')
    35000.0

    >>> sal(12345,69420,'BYR')
    977500.575
    """
    return currency_to_rub[sal_list[2]] * (float(sal_list[0]) + float(sal_list[1])) / 2


def for_loop_div(key_source: dict, divide: dict, action):
    """ Выполняет action от divide[x] и key_source[x] для каждого ключа из key_source


    :param key_source: Словарь с ключами
    :param divide: Словарь для заполнения
    :param action: lambda/def/method от двух агрументов
    """
    for x in key_source.keys():
        divide[x] = action(divide[x], key_source[x])


def csv_reader(file_name):
    """ Считывает данные из csv файла

    :param file_name: Имя файла с расщирением
    :return: list Список данных
    """
    file = codecs.open(file_name, 'r', 'utf_8_sig')
    reader = csv.reader(file)
    data = list(reader)
    file.close()
    return data


def addToDict(key_val, dict, val):
    """ Считает колво вхождений значения в словарь

    :param key_val: Ключ для словаря
    :param dict: Словарь - счётчик
    :param val: Значение для прибавления

    >>> temp_dict = {}
    >>> addToDict('a',temp_dict,1)
    >>> temp_dict
    {'a': 1}

    >>> addToDict('a', temp_dict,5)
    >>> addToDict('b', temp_dict,'gg')
    >>> temp_dict
    {'a': 6, 'b': 'gg'}
    """
    if key_val in dict.keys():
        dict[key_val] += val
    else:
        dict[key_val] = val


def create_dicts(data, name):
    """ Создаёт словари для заполнения Report класса

    :param data: Данные
    :param name: Имя профессии
    :return: int Общее число вакансий
    """
    number = 0
    for line in data:
        if all(line):
            number += 1
            addToDict(year(line), salary_all_years,
                      sal(line[Keys.salary_from], line[Keys.salary_to], line[Keys.salary_currency]))
            addToDict(line[Keys.area_name], salary_city,
                      sal(line[Keys.salary_from], line[Keys.salary_to], line[Keys.salary_currency]))
            addToDict(year(line), count_all_vacs, 1)
            addToDict(line[Keys.area_name], count_city_vacs, 1)
            if name in line[Keys.name]:
                addToDict(year(line), salary_prof_years,
                          sal(line[Keys.salary_from], line[Keys.salary_to], line[Keys.salary_currency]))
                addToDict(year(line), count_prof_vacs, 1)
    for_loop_div(count_all_vacs, salary_all_years, lambda x, y: int(x / y))
    for_loop_div(count_prof_vacs, salary_prof_years, lambda x, y: int(x / y))
    return number


def calculate_part_city(num):
    """ Процент от всего кол-ва вакансий по городам

    :param num: Число всех вакансий
    """
    for x in count_city_vacs.keys():
        calc_num = Decimal(count_city_vacs[x] / num).quantize(Decimal("1.0000"))
        if calc_num >= 0.01:
            part_city[x] = calc_num.__float__()
            salary_city_part[x] = int(salary_city[x] / count_city_vacs[x])


def alphabetic_sort(ls: list):
    """ Сортирует список кортежей по алфавиту и по значению в обратном порядке

    :param ls: Список
    """
    temp_list = []
    for i in range(1, len(ls)):
        if ls[i][1] == ls[i - 1][1]:
            temp_list.append(ls[i])
            if ls[i - 1] not in temp_list:
                temp_list.append(ls[i - 1])
        elif len(temp_list) > 1 and ls[i][1] != ls[i - 1][1]:
            ls[ls.index(temp_list[0]) - 1:i] = sorted(temp_list, key=lambda f: f[0])
            temp_list.clear()


file_name = input("Введите название файла: ")
prof_name = input("Введите название профессии: ")

data = csv_reader(file_name)
Keys = ProfKeys(data.pop(0))
number_of_vacs = create_dicts(data, prof_name)
calculate_part_city(number_of_vacs)
city_part_vacs = sorted(part_city.items(), key=lambda x: x[1], reverse=True)
city_salary = sorted(salary_city_part.items(), key=lambda x: x[1], reverse=True)
alphabetic_sort(city_salary)
fill(salary_prof_years, salary_all_years, 0)
fill(count_prof_vacs, count_all_vacs, 0)
report = Report(salary_all_years, count_all_vacs, salary_prof_years, count_prof_vacs, dict(city_salary[0:10]),
                dict(city_part_vacs[0:10]))
report.print_data()
report.generate_image()


if __name__ == '__main__':
    doctest.testmod()